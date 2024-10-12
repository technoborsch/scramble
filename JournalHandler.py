import sys
import os
import openpyxl

from config import JOURNAL_ADDRESS


def find_journal():
    if os.path.exists(JOURNAL_ADDRESS):
        return JOURNAL_ADDRESS
    else:
        journal_folder_path = "\\".join(JOURNAL_ADDRESS.split("\\")[:-1])
        journal_name = JOURNAL_ADDRESS.split("\\")[-1].split(".")[0]
        path = None
        if os.path.isdir(journal_folder_path):
            for file in os.listdir(journal_folder_path):
                if not file.startswith("~$") and journal_name in file:
                    this_path = journal_folder_path + "\\" + file
                    if path is None or os.path.getmtime(this_path) > os.path.getmtime(path):
                        path = this_path
        if path:
            return path
        elif hasattr(sys, "_MEIPASS"):
            return os.path.join(sys._MEIPASS, "journal.xlsx")
        else:
            return "materials\\journal.xlsx"



class JournalHandler:

    def __init__(self):
        self.path = find_journal()
        self.wb = openpyxl.open(self.path, read_only=True)
        self.ws = self.wb["Журнал ИИ"]

    def get_change_notice_info(self, change_notice_number):
        i = 1
        for cn_number in self.ws.iter_rows(1, self.ws.max_row, 5, 5, True):
            if cn_number[0] and str(change_notice_number) in str(cn_number[0]):
                return self.extract_info(i), i
            i += 1

    def extract_info(self, row):
        estimates = False
        if self.ws.cell(row, 12).value == "Требуется":
            estimates = True
        full_set_code = self.ws.cell(row, 7).value
        set_code, set_revision = full_set_code.split(",")[0].split("_")
        last_author = self.ws.cell(row, 17).value.split("/")[-1].strip(" ")
        change_number = self.ws.cell(row, 9).value
        release_date = self.ws.cell(row, 25).value
        return {
            "change_notice_number": self.ws.cell(row, 5).value,
            "full_set_code": full_set_code,
            "set_code": set_code,
            "set_revision": int(set_revision[1:]),
            "change_number": change_number,
            "release_date": release_date,
            "estimates": estimates,
            "last_author": last_author
        }

    def get_previous_change_notices_info(self, full_set_code, row):
        previous_cn = []
        j = 1
        for this_set_code in self.ws.iter_rows(1, row - 1, 7, 7, True):
            if this_set_code[0] and full_set_code in str(this_set_code[0]):
                previous_cn.append(self.extract_info(j))
            j += 1
        return previous_cn


if __name__ == "__main__":
    handler = JournalHandler()
    info, row = handler.get_change_notice_info(4808)
    print(info)
    previous = handler.get_previous_change_notices_info(info["full_set_code"], row)
    print(previous)
